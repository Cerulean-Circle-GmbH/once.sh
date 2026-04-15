#!/usr/bin/env python3
"""
AlmaLinux Work Report Generator
Generates a 5-sheet Excel workbook summarising AlmaLinux work (Feb–Apr 2026).
Run with: /tmp/xlsx-venv/bin/python3 docs/reports/generate_almalinux_report.py
"""

import os
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("openpyxl is required. Install with:")
    print("  python3 -m venv /tmp/xlsx-venv && /tmp/xlsx-venv/bin/pip install openpyxl")
    raise SystemExit(1)

# ---------------------------------------------------------------------------
# Color constants
# ---------------------------------------------------------------------------
DARK_BLUE       = "2F5496"
LIGHT_BLUE      = "D6E4F0"
MEDIUM_BLUE     = "B4C6E7"
DARK_BLUE_GRAD  = "8DB4E2"
LIGHT_GREY      = "F2F2F2"
GREEN           = "C6EFCE"
LIGHT_ORANGE    = "FCE4D6"
WHITE           = "FFFFFF"
GREY_FONT       = "808080"
DARK_GREY_FONT  = "404040"

# ---------------------------------------------------------------------------
# Shared style objects
# ---------------------------------------------------------------------------

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _side(style="thin", color="000000"):
    return Side(style=style, color=color)

def _thin_border():
    s = _side()
    return Border(left=s, right=s, top=s, bottom=s)

# Fonts
FONT_TITLE      = Font(name="Calibri", size=14, bold=True)
FONT_HEADER     = Font(name="Calibri", size=11, bold=True, color=WHITE)
FONT_META_LABEL = Font(name="Calibri", size=10, bold=True, color=DARK_BLUE)
FONT_META_VALUE = Font(name="Calibri", size=10, color="000000")
FONT_NORMAL     = Font(name="Calibri", size=10)
FONT_BOLD       = Font(name="Calibri", size=10, bold=True)
FONT_ITALIC     = Font(name="Calibri", size=11, italic=True)

# KPI fonts
FONT_KPI_VALUE  = Font(name="Calibri", size=28, bold=True, color=DARK_BLUE)
FONT_KPI_LABEL  = Font(name="Calibri", size=10, color=GREY_FONT)
FONT_KPI_DESC   = Font(name="Calibri", size=9,  color=DARK_GREY_FONT)

FONT_FOOTNOTE   = Font(name="Calibri", size=9, italic=True, color=GREY_FONT)

# Fills
FILL_DARK_BLUE  = _fill(DARK_BLUE)
FILL_LIGHT_BLUE = _fill(LIGHT_BLUE)
FILL_LIGHT_GREY = _fill(LIGHT_GREY)
FILL_GREEN      = _fill(GREEN)
FILL_LIGHT_ORG  = _fill(LIGHT_ORANGE)
FILL_WHITE      = _fill(WHITE)

# Alignments
ALIGN_CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
ALIGN_LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
ALIGN_LEFT_TOP  = Alignment(horizontal="left",   vertical="top",    wrap_text=True)
ALIGN_RIGHT     = Alignment(horizontal="right",  vertical="center")

# ---------------------------------------------------------------------------
# Helper functions
# ---------------------------------------------------------------------------

def apply_table_header(cell, text):
    """Style a header cell: dark blue fill, white bold font, centered."""
    cell.value     = text
    cell.font      = FONT_HEADER
    cell.fill      = FILL_DARK_BLUE
    cell.alignment = ALIGN_CENTER
    cell.border    = _thin_border()

def apply_data_cell(cell, value, fill=None, alignment=None, font=None, border=True):
    cell.value = value
    cell.font  = font or FONT_NORMAL
    if fill:
        cell.fill = fill
    cell.alignment = alignment or ALIGN_LEFT
    if border:
        cell.border = _thin_border()

def apply_bold_cell(cell, value, fill=None, alignment=None):
    apply_data_cell(cell, value, fill=fill, alignment=alignment, font=FONT_BOLD)

def set_col_widths(ws, widths_dict):
    """widths_dict: {col_letter_or_index: width}"""
    for col, width in widths_dict.items():
        if isinstance(col, int):
            col = get_column_letter(col)
        ws.column_dimensions[col].width = width

def row_height(ws, row, height):
    ws.row_dimensions[row].height = height

# ---------------------------------------------------------------------------
# 2. build_dashboard(wb)
# ---------------------------------------------------------------------------

def build_dashboard(wb):
    ws = wb.create_sheet("Dashboard")

    # Column widths — 10 columns, each ~12
    for i in range(1, 11):
        ws.column_dimensions[get_column_letter(i)].width = 13

    # --- Row 1: Title ---
    ws.merge_cells("A1:J1")
    c = ws["A1"]
    c.value     = "OOSH \u2014 AlmaLinux Platform Support"
    c.font      = FONT_TITLE
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_LIGHT_BLUE
    row_height(ws, 1, 32)

    # --- Rows 2-4: Metadata ---
    meta = [
        ("Project:", "OOSH \u2014 Object Oriented Shell"),
        ("Period:",  "2026-02-22 to 2026-04-13 (~7 weeks)"),
        ("Authors:", "Hannes Nortje, Marcel Donges"),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:J{i}")
        lc = ws.cell(row=i, column=1)
        lc.value     = label
        lc.font      = FONT_META_LABEL
        lc.alignment = ALIGN_LEFT
        vc = ws.cell(row=i, column=3)
        vc.value     = value
        vc.font      = FONT_META_VALUE
        vc.alignment = ALIGN_LEFT
        row_height(ws, i, 16)

    # Row 5 spacer
    row_height(ws, 5, 6)

    # --- Row 6: KPI section label ---
    ws.merge_cells("A6:J6")
    c = ws["A6"]
    c.value     = "Key Metrics"
    c.font      = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
    c.alignment = ALIGN_LEFT
    row_height(ws, 6, 18)

    # --- Rows 7-9: KPI boxes ---
    kpi_data = [
        (1,  3,  "250+",    "Commits",         "Total work volume"),
        (3,  5,  "7",       "Weeks",            "Feb 22 \u2013 Apr 13"),
        (5,  7,  "25+",     "Files Modified",   "Across OOSH core"),
        (7,  9,  "362/363", "Tests Passing",    "AlmaLinux 9"),
        (9, 11,  "5/5",     "Platforms Pass",   "Must-pass gate"),
    ]
    for start_col, end_col, value, label, desc in kpi_data:
        sc = get_column_letter(start_col)
        ec = get_column_letter(end_col - 1)
        ws.merge_cells(f"{sc}7:{ec}7")
        ws.merge_cells(f"{sc}8:{ec}8")
        ws.merge_cells(f"{sc}9:{ec}9")

        vc = ws[f"{sc}7"]
        vc.value     = value
        vc.font      = FONT_KPI_VALUE
        vc.fill      = FILL_LIGHT_BLUE
        vc.alignment = ALIGN_CENTER
        vc.border    = _thin_border()

        lc = ws[f"{sc}8"]
        lc.value     = label
        lc.font      = FONT_KPI_LABEL
        lc.fill      = FILL_LIGHT_BLUE
        lc.alignment = ALIGN_CENTER
        lc.border    = _thin_border()

        dc = ws[f"{sc}9"]
        dc.value     = desc
        dc.font      = FONT_KPI_DESC
        dc.fill      = FILL_LIGHT_BLUE
        dc.alignment = ALIGN_CENTER
        dc.border    = _thin_border()

    row_height(ws, 7, 42)
    row_height(ws, 8, 16)
    row_height(ws, 9, 14)

    # Row 10 spacer
    row_height(ws, 10, 6)

    # --- Row 11: Effort Breakdown header ---
    ws.merge_cells("A11:J11")
    c = ws["A11"]
    c.value     = "Effort Breakdown"
    c.font      = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
    c.alignment = ALIGN_LEFT
    row_height(ws, 11, 18)

    # --- Row 12: Table headers ---
    eff_headers = ["", "Core AlmaLinux", "Cross-Platform", "Total"]
    col_map = [1, 3, 6, 9]  # columns A, C, F, I (merged ranges)
    col_ends = [2, 5, 8, 10]

    for idx, (col_s, col_e, header) in enumerate(zip(col_map, col_ends, eff_headers)):
        sc = get_column_letter(col_s)
        ec = get_column_letter(col_e)
        if col_s != col_e:
            ws.merge_cells(f"{sc}12:{ec}12")
        c = ws[f"{sc}12"]
        if header:
            apply_table_header(c, header)
        else:
            c.fill   = FILL_DARK_BLUE
            c.border = _thin_border()

    row_height(ws, 12, 18)

    # --- Rows 13-14: Effort data ---
    effort_rows = [
        ("Commits",     "161+", "91+",  "250+"),
        ("Focus Areas", "8",    "10",   "18"),
        ("Total Repo",  "",     "",     "800"),
    ]
    for i, (label, core, cross, total) in enumerate(effort_rows, start=13):
        # label cell
        ws.merge_cells(f"A{i}:B{i}")
        apply_bold_cell(ws[f"A{i}"], label)

        # Core AlmaLinux (light blue)
        ws.merge_cells(f"C{i}:E{i}")
        apply_data_cell(ws[f"C{i}"], core, fill=FILL_LIGHT_BLUE, alignment=ALIGN_CENTER)

        # Cross-Platform (light grey)
        ws.merge_cells(f"F{i}:H{i}")
        apply_data_cell(ws[f"F{i}"], cross, fill=FILL_LIGHT_GREY, alignment=ALIGN_CENTER)

        # Total (bold)
        ws.merge_cells(f"I{i}:J{i}")
        apply_bold_cell(ws[f"I{i}"], total, alignment=ALIGN_CENTER)

        row_height(ws, i, 16)

    # Row 16 spacer
    row_height(ws, 16, 6)

    # --- Row 17: Outcome statement ---
    ws.merge_cells("A17:J17")
    c = ws["A17"]
    c.value = (
        "AlmaLinux 9 is a must-pass gate for production. "
        "No code reaches prod without passing AlmaLinux tests."
    )
    c.font      = FONT_ITALIC
    c.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    row_height(ws, 17, 36)

    ws.sheet_view.showGridLines = True


# ---------------------------------------------------------------------------
# 3. build_timeline(wb)
# ---------------------------------------------------------------------------

def build_timeline(wb):
    ws = wb.create_sheet("Timeline")

    set_col_widths(ws, {"A": 8, "B": 20, "C": 22, "D": 65, "E": 12})

    # Row 1: Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "Weekly Timeline \u2014 AlmaLinux Work (Feb 22 \u2013 Apr 13, 2026)"
    c.font      = FONT_TITLE
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_LIGHT_BLUE
    row_height(ws, 1, 28)

    # Row 2: spacer
    row_height(ws, 2, 6)

    # Row 3: Headers
    headers = ["Week", "Dates", "Focus Area", "Key Deliverables", "Commits"]
    for col, h in enumerate(headers, start=1):
        apply_table_header(ws.cell(row=3, column=col), h)
    row_height(ws, 3, 18)

    # Timeline data
    timeline_data = [
        (1,  "Feb 22 \u2013 Mar 1",  "Foundation",
         "odocker wrapper created, dnf/yum support, which \u2192 command -v, SSH ControlMaster fixes",
         82),
        (2,  "Mar 1 \u2013 8",        "Platform Setup",
         "AlmaLinux Docker image, platform matrix defined, RHEL install fixes, moreutils removed",
         70),
        (3,  "Mar 8 \u2013 15",       "Release Pipeline",
         "promote script + state machine, automated platform tests, v1.0.0 released, sparse array fix",
         132),
        (4,  "Mar 15 \u2013 22",      "macOS + Portability",
         "macOS CI pipeline, login shell/.bashrc fixes, portable sed -i, test suite hardening",
         177),
        (5,  "Mar 22 \u2013 29",      "Multi-User Install",
         "sudo PATH fix, UPG handling, auto-create user, cross-user install, .bash_profile portability",
         196),
        (6,  "Mar 29 \u2013 Apr 5",   "Docker & SSH",
         "Docker socket access, shared SSH config + deploy key, POSIX BRE regex fix, notests param, git config",
         130),
        (7,  "Apr 5 \u2013 13",       "Stabilisation",
         "Report preparation, minor framework improvements",
         13),
    ]

    def commit_fill(count):
        if count <= 50:
            return _fill(LIGHT_GREY)
        elif count <= 100:
            return _fill(LIGHT_BLUE)
        elif count <= 150:
            return _fill(MEDIUM_BLUE)
        else:
            return _fill(DARK_BLUE_GRAD)

    for row_idx, (week, dates, focus, deliverables, commits) in enumerate(timeline_data, start=4):
        r = row_idx
        apply_data_cell(ws.cell(r, 1), week,         alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(r, 2), dates)
        apply_data_cell(ws.cell(r, 3), focus,        font=FONT_BOLD)
        apply_data_cell(ws.cell(r, 4), deliverables, alignment=ALIGN_LEFT_TOP)
        apply_data_cell(ws.cell(r, 5), commits,      fill=commit_fill(commits),
                        alignment=ALIGN_CENTER)
        row_height(ws, r, 36)

    # Totals row
    total_row = 4 + len(timeline_data)
    apply_bold_cell(ws.cell(total_row, 1), "TOTAL", alignment=ALIGN_CENTER)
    apply_bold_cell(ws.cell(total_row, 2), "")
    apply_bold_cell(ws.cell(total_row, 3), "")
    apply_bold_cell(ws.cell(total_row, 4), "")
    apply_bold_cell(ws.cell(total_row, 5), 800, alignment=ALIGN_CENTER)
    row_height(ws, total_row, 16)

    ws.freeze_panes = "A4"


# ---------------------------------------------------------------------------
# 4. build_summary(wb)
# ---------------------------------------------------------------------------

def build_summary(wb):
    ws = wb.create_sheet("Summary")

    set_col_widths(ws, {"A": 30, "B": 12, "C": 12, "D": 16, "E": 14, "F": 14})

    # Row 1: Title
    ws.merge_cells("A1:F1")
    c = ws["A1"]
    c.value     = "OOSH \u2014 AlmaLinux Platform Support"
    c.font      = FONT_TITLE
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_LIGHT_BLUE
    row_height(ws, 1, 28)

    # Rows 2-5: Metadata
    meta = [
        ("Project:",  "OOSH \u2014 Object Oriented Shell"),
        ("Period:",   "2026-02-22 \u2013 2026-04-13 (~7 weeks)"),
        ("Authors:",  "Hannes Nortje, Marcel Donges"),
        ("Total Commits:", "250+"),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:F{i}")
        lc = ws.cell(row=i, column=1)
        lc.value     = label
        lc.font      = FONT_META_LABEL
        lc.alignment = ALIGN_LEFT
        vc = ws.cell(row=i, column=3)
        vc.value     = value
        vc.font      = FONT_META_VALUE
        vc.alignment = ALIGN_LEFT
        row_height(ws, i, 14)

    # Row 6: spacer
    row_height(ws, 6, 6)

    # Row 7: Table headers
    headers = ["Work Area", "Category", "Commits", "Files Changed", "Period", "Status"]
    for col, h in enumerate(headers, start=1):
        apply_table_header(ws.cell(7, col), h)
    row_height(ws, 7, 18)

    # Data rows
    # (name, category, commits, files, period)
    core_rows = [
        ("Package Manager (dnf/yum)",        "Core",    "1",    "5",    "Feb 2026"),
        ("POSIX Compatibility",               "Core",    "5",    "14",   "Feb\u2013Apr 2026"),
        ("Docker Infrastructure (odocker)",   "Core",    "68",   "3+",   "Feb\u2013Apr 2026"),
        ("AlmaLinux Docker Image",            "Core",    "1",    "1",    "Mar 2026"),
        ("Installation Fixes (ossh install)",  "Core",    "4",    "6",    "Mar 2026"),
        ("User Management (RHEL)",            "Core",    "1",    "1",    "Mar 2026"),
        ("Platform Test Infrastructure",      "Core",    "61",   "10+",  "Mar\u2013Apr 2026"),
        ("Release Pipeline",                  "Core",    "20+",  "10+",  "Mar 2026"),
    ]
    cross_rows = [
        ("Portable Shell (sed -i)",           "Cross-platform",    "1",    "2",    "Mar 2026"),
        ("Login Shell / .bashrc Fixes",       "Cross-platform",    "6",    "3+",   "Mar 2026"),
        ("Cross-Platform User Management",    "Cross-platform",    "6",    "3+",   "Mar 2026"),
        ("Multi-User Install (ossh)",         "Cross-platform",    "13",   "4+",   "Mar\u2013Apr 2026"),
        ("SSH / ControlMaster",               "Cross-platform",    "5",    "2+",   "Feb\u2013Mar 2026"),
        ("Shared SSH Config + Deploy Key",    "Cross-platform",    "9",    "3+",   "Apr 2026"),
        ("Docker Socket Access",              "Cross-platform",    "10",   "3+",   "Apr 2026"),
        ("Test Suite Framework",              "Cross-platform",    "7",    "3+",   "Feb\u2013Mar 2026"),
        ("State Machine Fixes",               "Cross-platform",    "4",    "4+",   "Mar 2026"),
        ("Branching Strategy + Promote",      "Cross-platform",    "30+",  "5+",   "Mar 2026"),
    ]

    all_rows = core_rows + cross_rows
    for i, (name, category, commits, files, period) in enumerate(all_rows, start=8):
        fill = FILL_LIGHT_BLUE if category == "Core" else FILL_LIGHT_GREY
        apply_data_cell(ws.cell(i, 1), name,     fill=fill, font=FONT_BOLD if category == "Core" else FONT_NORMAL)
        apply_data_cell(ws.cell(i, 2), category, fill=fill, alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 3), commits,  alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 4), files,    alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 5), period)
        apply_data_cell(ws.cell(i, 6), "Complete", fill=FILL_GREEN, alignment=ALIGN_CENTER)
        row_height(ws, i, 16)

    # Totals row
    total_row = 8 + len(all_rows)
    apply_bold_cell(ws.cell(total_row, 1), "TOTAL")
    apply_bold_cell(ws.cell(total_row, 2), "")
    apply_bold_cell(ws.cell(total_row, 3), "250+", alignment=ALIGN_CENTER)
    apply_bold_cell(ws.cell(total_row, 4), "25+",  alignment=ALIGN_CENTER)
    apply_bold_cell(ws.cell(total_row, 5), "")
    apply_bold_cell(ws.cell(total_row, 6), "")
    row_height(ws, total_row, 16)

    # Auto-filter and freeze
    ws.auto_filter.ref = f"A7:F{8 + len(all_rows)}"
    ws.freeze_panes    = "A8"


# ---------------------------------------------------------------------------
# 5. build_detail(wb)
# ---------------------------------------------------------------------------

def build_detail(wb):
    ws = wb.create_sheet("Work Areas Detail")

    set_col_widths(ws, {"A": 38, "B": 45, "C": 60, "D": 30, "E": 40})

    # Row 1: Title
    ws.merge_cells("A1:E1")
    c = ws["A1"]
    c.value     = "Work Areas \u2014 Detailed Breakdown"
    c.font      = FONT_TITLE
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_LIGHT_BLUE
    row_height(ws, 1, 28)

    # Row 2: spacer
    row_height(ws, 2, 6)

    # Row 3: Headers
    headers = ["Work Area", "Description", "What Was Done", "Key Commits", "Key Files"]
    for col, h in enumerate(headers, start=1):
        apply_table_header(ws.cell(3, col), h)
    row_height(ws, 3, 18)

    # Detail data rows (matching Summary order)
    # (work_area, description, what_was_done, key_commits, key_files, category)
    detail_rows = [
        # Core
        (
            "Package Manager (dnf/yum)",
            "OOSH only supported apt-get and brew; AlmaLinux requires dnf/yum",
            "Added dnf/yum detection to bootstrap (init/oosh), framework core (oo cmd install/update), completion system (c2.install.linux), Python venv setup for tests (otest), and user-facing install hints (claudeFlow). 5 files, +47/-6 lines",
            "d6de39a",
            "init/oosh, oo, ng/c2, otest, claudeFlow",
            "Core",
        ),
        (
            "POSIX Compatibility",
            "AlmaLinux minimal install lacks GNU tools; which, moreutils, GNU sed extensions caused failures",
            "Replaced which with command -v in 9 files (fixed 'sudo: command not found' on bootstrap); removed moreutils dependency by replacing errno with Python3 os.strerror(); fixed POSIX BRE regex \\? causing 'bad regex' in completion system; set git user identity in test fixtures for containers lacking global git config",
            "0f23296, dc7e701, 8f6784b, 848c755",
            "init/oosh, config, debug, loop, ng/c2, state, this, oo",
            "Core",
        ),
        (
            "Docker Infrastructure (odocker)",
            "No Docker wrapper existed; platform testing needed container lifecycle management",
            "Built odocker wrapper from scratch: 18 methods including build, run.sshd, reset, rebuild, clone, enter, exec, install (auto-detects OS for Docker CLI), workspace/container/image list, status, disk, prune. ~450 lines, 90+ test assertions, full tab completion, documentation at docs/odocker.md",
            "68 commits (Feb\u2013Apr)",
            "odocker, test/test.odocker, docs/odocker.md",
            "Core",
        ),
        (
            "AlmaLinux Docker Image",
            "No AlmaLinux container image existed for automated testing",
            "Created nakedAlma/9.sshd workspace and Dockerfile: almalinux:9 base + openssh-server + sudo + wget (229 MB). Built matching images for Ubuntu 24.04 (218 MB), Debian 12 (172 MB), Alpine 3.19 (15.7 MB)",
            "3f59cf8",
            "workspaces/nakedAlma/9.sshd/Dockerfile",
            "Core",
        ),
        (
            "Installation Fixes (ossh install)",
            "ossh install failed on RHEL/AlmaLinux: wrong sudo group, missing TTY, PATH override by secure_path",
            "Fixed wheel vs sudo group detection in user.create; added -tt (force TTY) for sudo password prompts over SSH; solved AlmaLinux secure_path overriding PATH with sudo bash -lc; added glob guards for mv *.public_key; state machine guards for missing dirs; suppressed bootstrap 'no OS found' noise",
            "c243b05, 16ec820, 7362817, 9c76f42",
            "ossh, user, oo, init/oosh",
            "Core",
        ),
        (
            "User Management (RHEL)",
            "RHEL/AlmaLinux auto-creates User Private Group matching username, causing useradd to fail",
            "RHEL auto-creates a User Private Group with same name as user; if a 'dev' group already exists, useradd fails. Fixed with -N (skip UPG creation) and -g (use existing group)",
            "3c65194",
            "user",
            "Core",
        ),
        (
            "Platform Test Infrastructure",
            "No automated way to test OOSH on AlmaLinux or any other platform",
            "Built 3 new methods: os platform.test <platform>, os platform.test.all, os platform.list. Automated flow per platform: odocker reset (fresh container) \u2192 SSH key push \u2192 ossh install \u2192 test.suite core 1 \u2192 evaluate results. Added notests parameter for install-only testing. 362/363 assertions pass on AlmaLinux",
            "5db7742, 61 commits total",
            "os, defaults/platforms.env, test/test.os",
            "Core",
        ),
        (
            "Release Pipeline",
            "No gated release process; AlmaLinux needed to be a promotion gate",
            "Created promote script with 15-state PROMOTE state machine: dev\u2192stage gated by core tests, stage\u2192prod gated by platform tests on all 5 must-pass targets. Resumable pipeline (failed promotions resume from failing step). Includes promote.status and promote.report. First release v1.0.0 on 2026-03-11",
            "20+ commits (Mar 2026)",
            "promote, test/test.promote, docs/promote.md, docs/branching.md",
            "Core",
        ),
        # Cross-platform
        (
            "Portable Shell (sed -i)",
            "GNU sed -i and BSD sed -i '' differ; macOS/AlmaLinux compatibility broken",
            "GNU sed -i (Linux) vs BSD sed -i '' (macOS) differ. Implemented uname -s platform detection wrapper so all in-place edits work on both. Also added dynamic root home detection (/root vs /var/root) and macOS shell detection",
            "eb76df4",
            "oo, ossh",
            "Cross-platform",
        ),
        (
            "Login Shell / .bashrc Fixes",
            "OOSH environment failed to load in login shells inside Docker containers on various platforms",
            "Created .bash_profile that sources .bashrc on login shells (SSH, Docker containers). Used ~ instead of $HOME for portability. Set bash as default login shell during remote install. Fixed root login shell on Alpine (ash \u2192 bash). Restored root shell setup lost during merge",
            "3f691ec, 70abd31, 9b396a3, c3ad938, d1abbbf, 737cbcd",
            "ossh, init/oosh",
            "Cross-platform",
        ),
        (
            "Cross-Platform User Management",
            "user.create failed on Debian (/usr/sbin not in PATH) and Alpine (adduser differences)",
            "Runtime detection of useradd vs adduser (not hardcoded per OS). Added /usr/sbin to PATH for Debian where useradd lives outside default PATH. Handled Alpine's adduser differences and existing group conflicts. usermod compatibility across all platforms",
            "6b368a4, bcfde06, 91ad7e8, 2b4063f, 075b4a3, ca098aa",
            "user, ossh, oo",
            "Cross-platform",
        ),
        (
            "Multi-User Install (ossh)",
            "ossh install only supported single-user; CI required root install + second dev user",
            "Extended ossh install for multi-user scenarios: auto-create user if missing, root-first platform test with second dev user, cross-user home access via sudo -u, targetHome for install.user.remote, skip bootstrap if oosh already installed. CI hardening: tolerate connection drops, local sudo for sudoers, fix awk escaping",
            "13 commits (Mar\u2013Apr)",
            "ossh, os",
            "Cross-platform",
        ),
        (
            "SSH / ControlMaster",
            "SSH session management caused hangs and stale connections during platform tests",
            "Set OSSH_CONTROL_PATH for sshpass and non-interactive ssh-keygen. Clean stale ControlMaster socket before sshpass connection. Use 'true' command instead of -N -f for ControlMaster. Reopen ControlMaster after install (group membership refresh). Close ControlMaster so dev group changes take effect",
            "97d8cce, f46df83, 2c70729, 9961051, d65708a",
            "ossh",
            "Cross-platform",
        ),
        (
            "Shared SSH Config + Deploy Key",
            "Multi-user Docker containers needed shared GitHub SSH access across users",
            "Designed and built shared SSH config for multi-user Docker containers: ossh.config.shared.create at /home/shared/.ssh/config, ossh.config.shared.link per-user, integrated into ossh.install. Copy (not symlink) SSH config per-user for robustness. Deploy key copied per-user with ~/.ssh/2cuGitHub identity",
            "fafbdd2, 59d6341, 1b5c427, 233c62d, ba85e25, 178bb0d, 11cb876, 8537ca9, 4790fae",
            "ossh",
            "Cross-platform",
        ),
        (
            "Docker Socket Access",
            "Containers (including AlmaLinux) needed Docker-in-Docker for platform testing",
            "Add users to docker group on install for socket access. Auto-detect and fix Docker socket GID mismatch (host vs container). Use sudo -n for non-interactive socket fix. Suppress oosh logging during group setup (logging not yet loaded). Check socket access only before build/rebuild/reset (performance)",
            "86744e3, 759168f, 420ba44, 3ef277c, dda4742, f6327a7, 9a39a65, 36dda9e, 87184ef, 41d0c0e",
            "odocker",
            "Cross-platform",
        ),
        (
            "Test Suite Framework",
            "Test suite failed across Docker root, test user, and AlmaLinux environments",
            "Split tests into core (22 files, gates promotion) and extended (27 files) categories. Fixed test.suite all 1 to pass on host, Docker root, and Docker test user. Use /tmp for temp files (permission fix for test users). Use $OOSH_DIR instead of relative paths. Fixed aggregation counters, debug toggle, and completion registration",
            "edbd6be, d956487, 974554a, 55979f2, f692d9c, 67acc90, 541105a",
            "test.suite, otest",
            "Cross-platform",
        ),
        (
            "State Machine Fixes",
            "Sparse array lookup caused state machine to skip states; found during v1.0.0 AlmaLinux promotion",
            "Fixed sparse array lookup in state machine advancement \u2014 states [21]-[26] were unreachable, discovered during v1.0.0 stage\u2192prod promotion. Comprehensive fixes to framework and all consumers. Persist state machine files across users. Set git safe.directory for multi-user installs. Added Darwin support",
            "162e026, dbd1431, fd4350d, 90cdaa4",
            "state, oo, user",
            "Cross-platform",
        ),
        (
            "Branching Strategy + Promote",
            "No formal release pipeline; AlmaLinux needed to be a mandatory gate before production",
            "Defined complete dev\u2192stage\u2192prod branching strategy. Built promote script with 15 state-machine check functions. Stage promotion gated by core tests. Production promotion gated by platform tests on all 5 must-pass targets (including AlmaLinux). Resumable pipeline \u2014 failed promotions restart from failing step. Skip parameter for manual override. promote.status and promote.report for visibility",
            "8e16e6b, e3e09ca, 5db7742, 4f74c87, a52b9c1 + 25 more",
            "promote, docs/branching.md, docs/promote.md",
            "Cross-platform",
        ),
    ]

    for i, (area, desc, done, commits, files, category) in enumerate(detail_rows, start=4):
        fill = FILL_LIGHT_BLUE if category == "Core" else FILL_LIGHT_GREY
        apply_data_cell(ws.cell(i, 1), area,    fill=fill, font=FONT_BOLD, alignment=ALIGN_LEFT_TOP)
        apply_data_cell(ws.cell(i, 2), desc,    alignment=ALIGN_LEFT_TOP)
        apply_data_cell(ws.cell(i, 3), done,    alignment=ALIGN_LEFT_TOP)
        apply_data_cell(ws.cell(i, 4), commits, alignment=ALIGN_LEFT_TOP,
                        font=Font(name="Calibri", size=9))
        apply_data_cell(ws.cell(i, 5), files,   alignment=ALIGN_LEFT_TOP,
                        font=Font(name="Calibri", size=9))
        row_height(ws, i, 48)

    last_data_row = 3 + len(detail_rows)
    ws.auto_filter.ref = f"A3:E{last_data_row}"
    ws.freeze_panes    = "A4"


# ---------------------------------------------------------------------------
# 6. build_platform_matrix(wb)
# ---------------------------------------------------------------------------

def build_platform_matrix(wb):
    ws = wb.create_sheet("Platform Matrix")

    set_col_widths(ws, {"A": 20, "B": 20, "C": 14, "D": 14, "E": 18, "F": 12, "G": 12})

    # Row 1: Title
    ws.merge_cells("A1:G1")
    c = ws["A1"]
    c.value     = "Platform Test Matrix \u2014 OOSH v1.0.20"
    c.font      = FONT_TITLE
    c.alignment = ALIGN_CENTER
    c.fill      = FILL_LIGHT_BLUE
    row_height(ws, 1, 28)

    # Rows 2-3: Metadata
    meta = [
        ("Test Date:", "2026-04-13"),
        ("Release:",   "v1.0.20"),
    ]
    for i, (label, value) in enumerate(meta, start=2):
        ws.merge_cells(f"A{i}:B{i}")
        ws.merge_cells(f"C{i}:G{i}")
        lc = ws.cell(i, 1)
        lc.value     = label
        lc.font      = FONT_META_LABEL
        lc.alignment = ALIGN_LEFT
        vc = ws.cell(i, 3)
        vc.value     = value
        vc.font      = FONT_META_VALUE
        vc.alignment = ALIGN_LEFT
        row_height(ws, i, 14)

    # Row 4: spacer
    row_height(ws, 4, 6)

    # Row 5: Headers
    headers = ["Platform", "Package Manager", "Tier", "Image Size", "Test Assertions", "Pass Rate", "Status"]
    for col, h in enumerate(headers, start=1):
        apply_table_header(ws.cell(5, col), h)
    row_height(ws, 5, 18)

    # Platform data rows
    # (platform, pkg_mgr, tier, image_size, assertions, pass_rate, status, is_almalinux)
    platform_data = [
        ("Ubuntu 24.04",  "apt-get", "must-pass",   "218 MB",  "362/363", "100%",  "Pass", False),
        ("Debian 12",     "apt-get", "must-pass",   "172 MB",  "362/363", "100%",  "Pass", False),
        ("AlmaLinux 9",   "dnf",     "must-pass",   "229 MB",  "362/363", "100%",  "Pass", True),
        ("Alpine 3.19",   "apk",     "must-pass",   "15.7 MB", "362/363", "100%",  "Pass", False),
        ("macOS",         "brew",    "must-pass",   "N/A",     "362/363", "100%",  "Pass", False),
        ("CentOS 7",      "yum",     "best-effort", "N/A",     "N/A",     "N/A",   "N/A",  False),
        ("Arch Linux",    "pacman",  "best-effort", "N/A",     "N/A",     "N/A",   "N/A",  False),
    ]

    for i, (platform, pkg, tier, size, assertions, rate, status, is_alma) in \
            enumerate(platform_data, start=6):
        tier_fill = FILL_LIGHT_ORG if tier == "must-pass" else FILL_LIGHT_GREY
        row_font  = FONT_BOLD if is_alma else FONT_NORMAL

        apply_data_cell(ws.cell(i, 1), platform,   font=row_font)
        apply_data_cell(ws.cell(i, 2), pkg,        alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 3), tier,       fill=tier_fill, alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 4), size,       alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 5), assertions, alignment=ALIGN_CENTER)
        apply_data_cell(ws.cell(i, 6), rate,       alignment=ALIGN_CENTER)

        if status == "Pass":
            apply_data_cell(ws.cell(i, 7), status, fill=FILL_GREEN, alignment=ALIGN_CENTER,
                            font=FONT_BOLD)
        else:
            apply_data_cell(ws.cell(i, 7), status, alignment=ALIGN_CENTER)

        row_height(ws, i, 16)

    # Blank spacer row before footnote
    blank_row = 6 + len(platform_data)
    row_height(ws, blank_row, 8)

    # Footnote row (row 14 when platform_data has 7 rows: blank=13, footnote=14)
    footnote_row = blank_row + 1
    ws.merge_cells(f"A{footnote_row}:G{footnote_row}")
    c = ws.cell(footnote_row, 1)
    c.value = (
        "* 1 intentional meta-test failure on AlmaLinux 9: "
        "a test that verifies failure detection itself \u2014 expected behaviour, not a regression."
    )
    c.font      = FONT_FOOTNOTE
    c.alignment = ALIGN_LEFT
    row_height(ws, footnote_row, 22)

    # Spacer
    spacer_row = footnote_row + 1
    row_height(ws, spacer_row, 8)

    # Promotion Gates sub-table (header at row 16)
    gate_title_row = spacer_row + 1
    ws.merge_cells(f"A{gate_title_row}:G{gate_title_row}")
    c = ws.cell(gate_title_row, 1)
    c.value     = "Promotion Gates"
    c.font      = Font(name="Calibri", size=12, bold=True, color=DARK_BLUE)
    c.alignment = ALIGN_LEFT
    row_height(ws, gate_title_row, 18)

    # Sub-table headers at row 17
    gate_header_row = gate_title_row + 1
    gate_headers = ["Gate", "Requirement", "Result"]
    col_spans = [(1, 2), (3, 5), (6, 7)]
    for (sc, ec), h in zip(col_spans, gate_headers):
        sl = get_column_letter(sc)
        el = get_column_letter(ec)
        if sc != ec:
            ws.merge_cells(f"{sl}{gate_header_row}:{el}{gate_header_row}")
        apply_table_header(ws.cell(gate_header_row, sc), h)
    row_height(ws, gate_header_row, 18)

    # Data at rows 18-19
    gates = [
        ("dev \u2192 stage", "All core tests pass",                           "Passed"),
        ("stage \u2192 prod", "Platform tests pass on all 5 must-pass targets (5/5)", "Passed 5/5"),
    ]
    for j, (gate, condition, result) in enumerate(gates, start=gate_header_row + 1):
        sl_a = get_column_letter(1)
        sl_c = get_column_letter(3)
        sl_f = get_column_letter(6)

        ws.merge_cells(f"{sl_a}{j}:{get_column_letter(2)}{j}")
        ws.merge_cells(f"{sl_c}{j}:{get_column_letter(5)}{j}")
        ws.merge_cells(f"{sl_f}{j}:{get_column_letter(7)}{j}")

        apply_data_cell(ws.cell(j, 1), gate,      font=FONT_BOLD)
        apply_data_cell(ws.cell(j, 3), condition)
        apply_data_cell(ws.cell(j, 6), result,    fill=FILL_GREEN, alignment=ALIGN_CENTER,
                        font=FONT_BOLD)
        row_height(ws, j, 18)

    ws.freeze_panes = "A6"


# ---------------------------------------------------------------------------
# 7. main()
# ---------------------------------------------------------------------------

def main():
    OUTPUT = os.path.join(
        os.path.dirname(os.path.abspath(__file__)),
        "2026-04-13-almalinux-work-report.xlsx"
    )

    wb = Workbook()
    # Remove the default blank sheet
    default_sheet = wb.active
    wb.remove(default_sheet)

    build_dashboard(wb)
    build_timeline(wb)
    build_summary(wb)
    build_detail(wb)
    build_platform_matrix(wb)

    wb.save(OUTPUT)

    print(f"Created: {OUTPUT}")
    print(f"Sheets:  {', '.join(ws.title for ws in wb.worksheets)}")
    size_kb = os.path.getsize(OUTPUT) / 1024
    print(f"Size:    {size_kb:.1f} KB")


if __name__ == "__main__":
    main()
